# from __future__ import annotations は必ずファイルの最上部に1回だけ記載
from __future__ import annotations

import argparse
import os
import shutil
import sys
import tempfile
import threading
import time
import traceback
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

import tkinter as tk
import tkinter.ttk as ttk
from tkinter import filedialog, messagebox, scrolledtext

from documentChecker import main

SUPPORTED_EXTENSIONS = (
    ".docx", ".doc", ".xlsx", ".xls", ".pptx", ".ppt", ".pdf", ".vsd", ".vsdx"
)
VISIO_EXTENSIONS = (".vsd", ".vsdx")
SUGGESTED_ACTION_SETTINGS_SHEET = "suggested_action_settings"


# ------------------------------
# 共通ユーティリティ
# ------------------------------
def build_visual_assets_dir(out_xlsx: str) -> str:
    output_path = Path(out_xlsx).resolve()
    return str((output_path.parent / f"{output_path.stem}_visual_assets").resolve())


def find_supported_files(folder: str) -> List[str]:
    files: List[str] = []
    for root, dirs, filenames in os.walk(folder):
        dirs[:] = [d for d in dirs if d not in {".git", "__pycache__", ".venv", "venv"}]
        for filename in filenames:
            if filename.startswith("~$"):
                continue
            if filename.lower().endswith(SUPPORTED_EXTENSIONS):
                files.append(os.path.join(root, filename))
    return files


def format_seconds(seconds: float) -> str:
    try:
        seconds = float(seconds)
    except Exception:
        seconds = 0.0
    if seconds >= 60:
        minutes = int(seconds // 60)
        remain = seconds - (minutes * 60)
        return f"{minutes}分{remain:.1f}秒"
    return f"{seconds:.1f}秒"


def _copy_file_preserve_tree(src: Path, dst_root: Path, src_root: Path) -> Path:
    rel = src.resolve().relative_to(src_root.resolve())
    dst = dst_root / rel
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)
    return dst


# ------------------------------
# Visio 前処理
# ------------------------------

def _runtime_base_dirs() -> List[Path]:
    """Windows配布向けに、実行ファイル周辺とPyInstaller展開先を探索する。"""
    dirs: List[Path] = []
    try:
        if getattr(sys, "frozen", False):
            dirs.append(Path(sys.executable).resolve().parent)
        dirs.append(Path(__file__).resolve().parent)
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            dirs.append(Path(meipass).resolve())
    except Exception:
        pass

    unique: List[Path] = []
    seen = set()
    for d in dirs:
        key = str(d).lower()
        if key not in seen:
            unique.append(d)
            seen.add(key)
    return unique


def _add_aspose_plugin_paths() -> List[str]:
    candidates: List[Path] = []
    env_dir = os.environ.get("ASPOSE_DIAGRAM_PLUGIN_DIR", "").strip()
    if env_dir:
        candidates.append(Path(env_dir))

    for base_dir in _runtime_base_dirs():
        candidates.extend([
            base_dir / "plugins" / "aspose_diagram",
            base_dir / "plugins" / "aspose",
            base_dir / "plugins",
            base_dir / "vendor" / "aspose_diagram",
            base_dir / "vendor" / "aspose",
            base_dir / "vendor",
            base_dir / "lib" / "aspose_diagram",
            base_dir / "lib",
            base_dir / "_internal" / "plugins" / "aspose_diagram",
            base_dir / "_internal" / "plugins",
        ])

    added: List[str] = []
    for path in candidates:
        try:
            if path.exists() and path.is_dir():
                path_str = str(path.resolve())
                if path_str not in sys.path:
                    sys.path.insert(0, path_str)
                    added.append(path_str)
        except Exception:
            continue
    return added


def _ensure_aspose_diagram_available() -> Tuple[bool, str]:
    """
    オフラインWindows配布向け。pip自動導入は行わず、plugins配下の
    Aspose.Diagramプラグインを探索して読み込む。
    """
    try:
        import aspose.diagram  # type: ignore
        return True, "Aspose.Diagram 利用可能"
    except Exception as first_exc:
        added = _add_aspose_plugin_paths()
        try:
            import importlib
            importlib.invalidate_caches()
            import aspose.diagram  # type: ignore
            return True, "Aspose.Diagram プラグインを読み込みました。"
        except Exception as plugin_exc:
            hint = (
                "Aspose.Diagram プラグイン未配置または読み込み不可。"
                "exeと同じフォルダ配下の plugins\\aspose_diagram に aspose パッケージ一式を配置してください。"
            )
            if added:
                hint += f" 探索済み={'; '.join(added[:3])}"
            return False, f"{hint} import_error={first_exc}; plugin_error={plugin_exc}"


def _convert_vsd_to_vsdx(src: Path, dst: Path) -> Optional[str]:
    pythoncom = None
    visio = None
    try:
        import pythoncom as imported_pythoncom
        import win32com.client as win32_client

        pythoncom = imported_pythoncom
        pythoncom.CoInitialize()
        visio = win32_client.Dispatch("Visio.Application")
        visio.Visible = False
        visio.AlertResponse = 7

        document = None
        try:
            document = visio.Documents.Open(str(src))
            dst.parent.mkdir(parents=True, exist_ok=True)
            document.SaveAs(str(dst))
            return None
        finally:
            if document is not None:
                try:
                    document.Close()
                except Exception:
                    pass
    except Exception as exc_com:
        try:
            ok, aspose_msg = _ensure_aspose_diagram_available()
            if not ok:
                raise RuntimeError(aspose_msg)
            from aspose.diagram import Diagram, SaveFileFormat

            dst.parent.mkdir(parents=True, exist_ok=True)
            diagram = Diagram(str(src))
            diagram.save(str(dst), SaveFileFormat.VSDX)
            return None
        except Exception as exc_aspose:
            return f"VSD→VSDX変換失敗(COM={exc_com}; Aspose={exc_aspose})"
    finally:
        if visio is not None:
            try:
                visio.Quit()
            except Exception:
                pass
        if pythoncom is not None:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass


def _find_soffice_path() -> Optional[str]:
    for candidate in [
        os.environ.get("SOFFICE_PATH", ""),
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]:
        if candidate and os.path.exists(candidate):
            return candidate
    return None


def _convert_visio_to_pdf(src: Path, dst_pdf: Path) -> Optional[str]:
    pythoncom = None
    visio = None
    try:
        import pythoncom as imported_pythoncom
        import win32com.client as win32_client

        pythoncom = imported_pythoncom
        pythoncom.CoInitialize()
        visio = win32_client.Dispatch("Visio.Application")
        visio.Visible = False
        visio.AlertResponse = 7

        document = None
        try:
            document = visio.Documents.Open(str(src))
            dst_pdf.parent.mkdir(parents=True, exist_ok=True)
            # 1 = visFixedFormatPDF
            document.ExportAsFixedFormat(1, str(dst_pdf), 1, 0)
            return None
        finally:
            if document is not None:
                try:
                    document.Close()
                except Exception:
                    pass
    except Exception as exc_com:
        try:
            ok, aspose_msg = _ensure_aspose_diagram_available()
            if not ok:
                raise RuntimeError(aspose_msg)
            from aspose.diagram import Diagram, SaveFileFormat

            dst_pdf.parent.mkdir(parents=True, exist_ok=True)
            diagram = Diagram(str(src))
            diagram.save(str(dst_pdf), SaveFileFormat.PDF)
            return None
        except Exception as exc_aspose:
            try:
                import subprocess

                soffice = _find_soffice_path()
                if not soffice:
                    return f"Visio→PDF変換失敗(COM={exc_com}; Aspose={exc_aspose}; LibreOffice=not found)"
                outdir = dst_pdf.parent
                outdir.mkdir(parents=True, exist_ok=True)
                proc = subprocess.run(
                    [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(outdir), str(src)],
                    capture_output=True,
                    text=True,
                    timeout=180,
                    check=False,
                )
                generated = outdir / f"{src.stem}.pdf"
                if proc.returncode == 0 and generated.exists():
                    if generated.resolve() != dst_pdf.resolve():
                        shutil.move(str(generated), str(dst_pdf))
                    return None
                return (
                    f"Visio→PDF変換失敗(COM={exc_com}; Aspose={exc_aspose}; "
                    f"LibreOffice rc={proc.returncode} stdout={proc.stdout} stderr={proc.stderr})"
                )
            except Exception as exc_lo:
                return f"Visio→PDF変換失敗(COM={exc_com}; Aspose={exc_aspose}; LibreOffice={exc_lo})"
    finally:
        if visio is not None:
            try:
                visio.Quit()
            except Exception:
                pass
        if pythoncom is not None:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass


def stage_folder_pdf_first(
    folder: str,
    files: List[str],
) -> Tuple[Path, Optional[Path], List[str], Dict[str, Dict[str, str]]]:
    """
    安全優先のステージング:
      - 非Visioファイルはそのままコピー
      - .vsd は可能なら .vsdx へ変換
      - Visio(.vsd/.vsdx) は可能なら PDF プロキシへ変換
      - 以降の documentChecker.main() では PDF を軸に処理させる
      - 出力xlsx補正用に、PDFプロキシ→元Visio のマッピングを返す
    """
    root_path = Path(folder).resolve()
    stage_root = Path(tempfile.mkdtemp(prefix="documentChecker_pdf_stage_"))
    staged_root = stage_root / root_path.name
    staged_root.mkdir(parents=True, exist_ok=True)

    failure_lines: List[str] = []
    proxy_map: Dict[str, Dict[str, str]] = {}
    visio_sources: List[Path] = []

    # documentChecker.main() はステージングフォルダを走査するため、
    # VSD前処理時も対象外ファイルが other_files シートに出るように、
    # サポート対象外ファイルも一時フォルダへコピーする。
    ignored_dir_names = {".git", "__pycache__", ".venv", "venv"}
    for dirpath, dirnames, filenames in os.walk(root_path):
        dirnames[:] = [d for d in dirnames if d not in ignored_dir_names]
        dir_path = Path(dirpath)
        for filename in filenames:
            if filename.startswith("~$"):
                continue
            src = (dir_path / filename).resolve()
            if src.suffix.lower() in SUPPORTED_EXTENSIONS:
                continue
            try:
                _copy_file_preserve_tree(src, staged_root, root_path)
            except Exception:
                pass

    for src_str in files:
        src = Path(src_str).resolve()
        if src.suffix.lower() in VISIO_EXTENSIONS:
            visio_sources.append(src)
        else:
            _copy_file_preserve_tree(src, staged_root, root_path)

    if not visio_sources:
        return staged_root, None, failure_lines, proxy_map

    for src in visio_sources:
        rel = src.relative_to(root_path)
        working_src = src

        if src.suffix.lower() == ".vsd":
            vsdx_in_stage = staged_root / rel.with_suffix(".vsdx")
            err = _convert_vsd_to_vsdx(src, vsdx_in_stage)
            if err is None and vsdx_in_stage.exists():
                working_src = vsdx_in_stage
            else:
                failure_lines.append(f"WARN  {rel} -> {err or 'VSD→VSDX変換失敗'}")

        pdf_proxy = staged_root / rel.with_suffix(".pdf")
        pdf_err = _convert_visio_to_pdf(working_src, pdf_proxy)
        if pdf_err is None and pdf_proxy.exists():
            proxy_map[str(pdf_proxy.resolve())] = {
                "original_path": str(src.resolve()),
                "original_type": src.suffix.replace(".", "").upper(),
            }
        else:
            fallback_src = src if working_src == src else working_src
            try:
                if fallback_src.resolve().is_relative_to(staged_root.resolve()):
                    pass
                else:
                    _copy_file_preserve_tree(fallback_src, staged_root, root_path)
            except Exception:
                pass
            failure_lines.append(f"WARN  {rel} -> {pdf_err or 'Visio→PDF変換失敗'}")

    failure_log_path = None
    return staged_root, failure_log_path, failure_lines, proxy_map


# ------------------------------
# 出力xlsx補正
# ------------------------------
def rewrite_output_xlsx(
    out_xlsx: str,
    staged_root: Path,
    original_root: Path,
    proxy_map: Dict[str, Dict[str, str]],
) -> Tuple[int, int, bool]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return 0, 0, False

    try:
        wb = load_workbook(out_xlsx)
        replaced_cells = 0
        filetype_fixed = 0
        staged_prefix = str(staged_root.resolve())
        original_prefix = str(original_root.resolve())

        # 1) 全セルの temp パスを元フォルダパスへ置換
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str):
                        if value in proxy_map:
                            original_path = proxy_map[value]["original_path"]
                            if value != original_path:
                                cell.value = original_path
                                replaced_cells += 1
                        elif value.startswith(staged_prefix):
                            cell.value = original_prefix + value[len(staged_prefix):]
                            replaced_cells += 1

        # 2) file_path / file_type 列をヘッダ名ベースで補正（VisioプロキシPDF → 元VSD/VSDX）
        for ws in wb.worksheets:
            if ws.max_row < 2:
                continue
            headers = {}
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if isinstance(header, str):
                    headers[header.strip().lower()] = col
            file_path_col = headers.get("file_path")
            file_type_col = headers.get("file_type")
            if not file_path_col:
                continue
            for row_idx in range(2, ws.max_row + 1):
                file_path_cell = ws.cell(row=row_idx, column=file_path_col)
                file_path_value = file_path_cell.value
                if not isinstance(file_path_value, str):
                    continue
                # すでに 1) で original_path へ変わっている可能性があるため、逆引きも見る
                matched = None
                for staged_proxy, meta in proxy_map.items():
                    if file_path_value == meta["original_path"] or file_path_value == staged_proxy:
                        matched = meta
                        break
                if not matched:
                    continue
                file_path_cell.value = matched["original_path"]
                if file_type_col:
                    type_cell = ws.cell(row=row_idx, column=file_type_col)
                    if type_cell.value != matched["original_type"]:
                        type_cell.value = matched["original_type"]
                        filetype_fixed += 1

        # 3) suggested_action_settings を非表示
        hidden = False
        if SUGGESTED_ACTION_SETTINGS_SHEET in wb.sheetnames:
            ws = wb[SUGGESTED_ACTION_SETTINGS_SHEET]
            ws.sheet_state = "hidden"
            hidden = True

        wb.save(out_xlsx)
        return replaced_cells, filetype_fixed, hidden
    except Exception:
        return 0, 0, False


# ------------------------------
# 本体呼び出し
# ------------------------------
def real_check(
    folder: str,
    files: List[str],
    out_xlsx: str,
    action_config_xlsx: str,
    log_callback: Callable[[str], None],
    progress_callback: Callable[[Dict[str, object]], None],
    result_callback: Callable[[bool, str], None],
    cancel_flag: List[bool],
) -> None:
    stage_root: Optional[Path] = None
    visual_assets_dir = build_visual_assets_dir(out_xlsx)
    start_time = time.time()
    proxy_map: Dict[str, Dict[str, str]] = {}

    def is_cancelled() -> bool:
        return bool(cancel_flag and cancel_flag[0])

    try:
        if is_cancelled():
            result_callback(False, "キャンセルされました")
            return

        working_folder = folder
        visio_files = [p for p in files if p.lower().endswith(VISIO_EXTENSIONS)]
        if visio_files:
            stage_root, failure_log_path, failure_lines, proxy_map = stage_folder_pdf_first(folder, files)
            working_folder = str(stage_root)

        argv = [working_folder, "-o", out_xlsx, "--visual-assets-dir", visual_assets_dir]
        if action_config_xlsx:
            argv += ["--action-config-xlsx", action_config_xlsx]

        def bridge_progress(payload: Dict[str, object]) -> None:
            progress_callback(payload)

        main(argv, progress_callback=bridge_progress, cancel_requested=is_cancelled)

        if is_cancelled():
            result_callback(False, "キャンセルされました")
            return

        if stage_root is not None and Path(out_xlsx).exists():
            rewrite_output_xlsx(
                out_xlsx=out_xlsx,
                staged_root=stage_root,
                original_root=Path(folder).resolve(),
                proxy_map=proxy_map,
            )
        elapsed = time.time() - start_time
        result_callback(True, f"チェックが完了しました|{elapsed}")

    except Exception as exc:
        log_callback(f"例外が発生しました: {exc}")
        log_callback(traceback.format_exc())
        result_callback(False, f"例外が発生しました: {exc}")
    finally:
        if stage_root is not None:
            shutil.rmtree(stage_root.parent, ignore_errors=True)


# ------------------------------
# GUI
# ------------------------------
class DocumentCheckerGUI(tk.Tk):
    def __init__(
        self,
        startup_folder: str = "",
        startup_output: str = "",
        startup_action_config: str = "",
        auto_start: bool = False,
        auto_close: bool = False,
    ):
        super().__init__()
        self.title("documentChecker GUI")
        self.geometry("820x500")
        self.minsize(760, 420)
        self.auto_close = auto_close
        self.check_thread: Optional[threading.Thread] = None
        self.cancel_flag = [False]
        self._finished = False
        self._run_start_ts = 0.0
        self._total_files = 0
        self._error_files: List[str] = []
        self._seen_processed: set[Tuple[int, str]] = set()
        self.create_widgets()
        if startup_folder:
            self.folder_var.set(startup_folder)
        if startup_output:
            self.outfile_var.set(startup_output)
        if startup_action_config:
            self.action_config_var.set(startup_action_config)
        if auto_start:
            self.after(300, self.run_check)

    def create_widgets(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        container = ttk.Frame(self, padding=(12, 10, 12, 10))
        container.grid(row=0, column=0, sticky="nsew")
        container.columnconfigure(1, weight=1)
        container.rowconfigure(4, weight=1)

        ttk.Label(container, text="対象フォルダ").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=(0, 8))
        self.folder_var = tk.StringVar()
        ttk.Entry(container, textvariable=self.folder_var).grid(row=0, column=1, sticky="ew", pady=(0, 8))
        ttk.Button(container, text="参照", command=self.browse_folder).grid(row=0, column=2, sticky="ew", padx=(8, 0), pady=(0, 8))

        ttk.Label(container, text="出力ファイル (xlsx)").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=(0, 8))
        self.outfile_var = tk.StringVar()
        ttk.Entry(container, textvariable=self.outfile_var).grid(row=1, column=1, sticky="ew", pady=(0, 8))
        ttk.Button(container, text="保存先", command=self.save_file).grid(row=1, column=2, sticky="ew", padx=(8, 0), pady=(0, 8))

        ttk.Label(container, text="対応推奨設定 (xlsx, 任意)").grid(row=2, column=0, sticky="w", padx=(0, 8), pady=(0, 8))
        self.action_config_var = tk.StringVar()
        ttk.Entry(container, textvariable=self.action_config_var).grid(row=2, column=1, sticky="ew", pady=(0, 8))
        ttk.Button(container, text="参照", command=self.browse_action_config).grid(row=2, column=2, sticky="ew", padx=(8, 0), pady=(0, 8))

        self.progress_var = tk.DoubleVar(value=0)
        self.progress = ttk.Progressbar(container, variable=self.progress_var, maximum=100)
        self.progress.grid(row=3, column=0, columnspan=3, sticky="ew", pady=(0, 8))

        self.log_text = scrolledtext.ScrolledText(container, height=16, width=90, state="disabled")
        self.log_text.grid(row=4, column=0, columnspan=3, sticky="nsew", pady=(0, 10))

        button_frame = ttk.Frame(container)
        button_frame.grid(row=5, column=0, columnspan=3, sticky="e")
        self.run_btn = ttk.Button(button_frame, text="実行", command=self.run_check)
        self.run_btn.grid(row=0, column=0, sticky="ew")
        self.stop_btn = ttk.Button(button_frame, text="中止", command=self.stop_check, state="disabled")
        self.stop_btn.grid(row=0, column=1, sticky="ew", padx=(8, 0))
        self.exit_btn = ttk.Button(button_frame, text="終了", command=self.destroy)
        self.exit_btn.grid(row=0, column=2, sticky="ew", padx=(8, 0))

    def browse_folder(self) -> None:
        folder = filedialog.askdirectory()
        if folder:
            self.folder_var.set(folder)

    def save_file(self) -> None:
        file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file:
            self.outfile_var.set(file)

    def browse_action_config(self) -> None:
        file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file:
            self.action_config_var.set(file)

    def log(self, msg: str) -> None:
        if threading.current_thread() is not threading.main_thread():
            self.after(0, self.log, msg)
            return
        self.log_text.config(state="normal")
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.log_text.config(state="disabled")
        self.update_idletasks()

    def update_progress(self, payload_or_idx, total: Optional[int] = None) -> None:
        if threading.current_thread() is not threading.main_thread():
            self.after(0, self.update_progress, payload_or_idx, total)
            return

        if isinstance(payload_or_idx, dict):
            payload = payload_or_idx
            total_files = int(payload.get("total", 0) or 0)
            completed = int(payload.get("completed", 0) or 0)
            processing = int(payload.get("processing", 0) or 0)
            current = completed + processing
            if total_files > 0:
                self.progress_var.set(max(0.0, min(100.0, 100.0 * current / total_files)))
            else:
                self.progress_var.set(0.0)

            phase = str(payload.get("phase", "") or "")
            if phase == "start":
                self._total_files = int(payload.get("all_count", total_files) or total_files or 0)
                self.log(f"処理開始　対象全体 {self._total_files}件")
            elif phase == "processed":
                file_index = int(payload.get("file_index", completed) or completed or 0)
                current_file = str(payload.get("current_file", "") or "")
                file_name = Path(current_file).name if current_file else "(不明)"
                elapsed_sec = float(payload.get("file_elapsed_sec", 0.0) or 0.0)
                file_result = str(payload.get("file_result", "") or "").lower()
                dedup_key = (file_index, file_name)
                if dedup_key not in self._seen_processed:
                    self._seen_processed.add(dedup_key)
                    self.log(f"[{file_index}/{self._total_files or total_files}] {file_name}　処理時間 {format_seconds(elapsed_sec)}")
                if file_result in {"ng", "error"}:
                    if file_name not in self._error_files:
                        self._error_files.append(file_name)
            elif phase == "done":
                self.progress_var.set(100.0)
        else:
            total_files = int(total or 0)
            current = int(payload_or_idx or 0)
            if total_files > 0:
                self.progress_var.set(max(0.0, min(100.0, 100.0 * current / total_files)))
            else:
                self.progress_var.set(0.0)

        self.update_idletasks()

    def _resolve_default_action_config(self) -> str:
        action_config_xlsx = self.action_config_var.get().strip()
        if action_config_xlsx:
            return action_config_xlsx
        if hasattr(sys, "_MEIPASS"):
            base_dir = sys._MEIPASS  # type: ignore[attr-defined]
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
        default_catalog = os.path.join(base_dir, "suggested_action_settings_catalog.xlsx")
        if os.path.isfile(default_catalog):
            self.action_config_var.set(default_catalog)
            return default_catalog
        return ""

    def run_check(self) -> None:
        try:
            if self.check_thread is not None and self.check_thread.is_alive():
                return

            self.log_text.config(state="normal")
            self.log_text.delete("1.0", "end")
            self.log_text.config(state="disabled")
            self.progress_var.set(0)
            self._finished = False
            self.cancel_flag = [False]
            self._run_start_ts = time.time()
            self._total_files = 0
            self._error_files = []
            self._seen_processed = set()

            folder = self.folder_var.get().strip()
            out_xlsx = self.outfile_var.get().strip()
            action_config_xlsx = self._resolve_default_action_config()

            if not folder or not os.path.isdir(folder):
                messagebox.showerror("エラー", "有効な対象フォルダを指定してください")
                return
            if not out_xlsx:
                messagebox.showerror("エラー", "出力ファイル(xlsx)を指定してください")
                return
            if action_config_xlsx and not os.path.isfile(action_config_xlsx):
                messagebox.showerror("エラー", "対応推奨設定(xlsx)が見つかりません")
                return

            files = find_supported_files(folder)
            if not files:
                messagebox.showinfo("情報", "対象ファイルが見つかりません")
                return

            # --- 追加: 処理開始時に待機メッセージを表示 ---
            self.log("準備中です。しばらくお待ちください。\n（ファイル数や環境によっては数分かかる場合があります）")

            self.run_btn.config(state="disabled")
            self.stop_btn.config(state="normal")

            self.check_thread = threading.Thread(
                target=real_check,
                args=(
                    folder,
                    files,
                    out_xlsx,
                    action_config_xlsx,
                    self.log,
                    self.update_progress,
                    self.check_done,
                    self.cancel_flag,
                ),
                daemon=True,
            )
            self.check_thread.start()
        except Exception as exc:
            self.log(f"例外が発生しました: {exc}")
            self.log(traceback.format_exc())
            messagebox.showerror("例外", f"run_check で例外が発生しました:\n{exc}")
            self.run_btn.config(state="normal")
            self.stop_btn.config(state="disabled")

    def stop_check(self) -> None:
        self.cancel_flag[0] = True
        self.log("キャンセル要求を受け付けました")

    def _build_summary_text(self, success: bool, message: str) -> str:
        elapsed_sec = time.time() - self._run_start_ts if self._run_start_ts else 0.0
        summary_lines: List[str] = []
        if success:
            summary_lines.append("正常に終了しました。")
            summary_lines.append(f"全体の処理時間 {format_seconds(elapsed_sec)}")
            summary_lines.append(f"エラー：{len(self._error_files)}件")
            for file_name in self._error_files:
                summary_lines.append(file_name)
        else:
            summary_lines.append(message)
            if message != "キャンセルされました":
                summary_lines.append(f"全体の処理時間 {format_seconds(elapsed_sec)}")
        return "\n".join(summary_lines)

    def check_done(self, success: bool, message: str) -> None:
        if threading.current_thread() is not threading.main_thread():
            self.after(0, self.check_done, success, message)
            return
        if self._finished:
            return
        self._finished = True
        self.run_btn.config(state="normal")
        self.stop_btn.config(state="disabled")

        # real_check からのメッセージ形式: "チェックが完了しました|<elapsed>"
        if success and "|" in message:
            _, _, raw_elapsed = message.partition("|")
            try:
                # GUIの計測と大差ないが、内部計測を優先したい場合は保持だけしておく
                _ = float(raw_elapsed)
            except Exception:
                pass

        summary_text = self._build_summary_text(success, message.split("|", 1)[0])
        self.log(summary_text)

        if success:
            messagebox.showinfo("完了", summary_text)
            if self.auto_close:
                self.after(300, self.destroy)
        else:
            if message == "キャンセルされました":
                messagebox.showinfo("キャンセル", summary_text)
            else:
                messagebox.showwarning("終了", summary_text)


# ------------------------------
# 起動引数
# ------------------------------
def parse_gui_args(argv: Optional[List[str]] = None):
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("--folder")
    parser.add_argument("--output")
    parser.add_argument("--action-config")
    parser.add_argument("--auto-start", action="store_true")
    parser.add_argument("--auto-close", action="store_true")
    args, _ = parser.parse_known_args(argv)

    output = args.output
    if args.auto_start and not output:
        output = str((Path.cwd() / "review_results_gui_new.xlsx").resolve())

    return args.folder or "", output or "", args.action_config or "", args.auto_start, args.auto_close


if __name__ == "__main__":
    startup_folder, startup_output, startup_action_config, auto_start, auto_close = parse_gui_args(sys.argv[1:])
    app = DocumentCheckerGUI(
        startup_folder=startup_folder,
        startup_output=startup_output,
        startup_action_config=startup_action_config,
        auto_start=auto_start,
        auto_close=auto_close,
    )
    app.mainloop()
